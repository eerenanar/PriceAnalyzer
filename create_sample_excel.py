"""
Örnek products.xlsx dosyası oluşturur.
Kullanım: python create_sample_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ürünler"

# Başlık
ws["A1"] = "Ürün Kodu"
header_cell = ws["A1"]
header_cell.font = Font(bold=True, color="FFFFFF")
header_cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
header_cell.alignment = Alignment(horizontal="center")

# Örnek ürün kodları
sample_codes = [
    "PRD-001",
    "PRD-002",
    "PRD-003",
    "ABC-123",
    "XYZ-456",
    "TEST-001",
    "SAMPLE-A",
    "SAMPLE-B",
]

for i, code in enumerate(sample_codes, start=2):
    ws[f"A{i}"] = code

ws.column_dimensions["A"].width = 20

output = "products.xlsx"
wb.save(output)
print(f"Örnek Excel oluşturuldu: {output}")
print(f"{len(sample_codes)} ürün kodu eklendi.")
