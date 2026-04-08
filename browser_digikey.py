"""Browser (Playwright) üzerinden Digikey fiyat/stok tarama scripti."""

import json
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout


def read_products(file_path: str):
    """oguz.xlsx'den ürün kodu ve adet bilgilerini okur."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(values_only=True):
        if row[0]:
            code = str(row[0]).strip()
            qty = int(row[1]) if row[1] else 0
            products.append((code, qty))
    wb.close()
    return products


def parse_price_from_text(text: str):
    """Fiyat metninden float değer çıkarır."""
    if not text:
        return None
    text = text.strip()
    match = re.search(r'\$?([\d,]+\.?\d*)', text)
    if match:
        return float(match.group(1).replace(',', ''))
    return None


def scrape_digikey_product(page, product_code: str, qty: int):
    """Tek bir ürünü Digikey'de arar ve fiyat/stok bilgisini döndürür."""
    url = f"https://www.digikey.com/en/products/result?keywords={product_code}"
    result = {
        'code': product_code,
        'qty': qty,
        'unit_price': None,
        'qty_price': None,
        'stock': None,
        'currency': 'USD',
        'url': url,
        'status': 'ERROR',
        'description': '',
    }

    try:
        page.goto(url, wait_until='domcontentloaded', timeout=30000)
        time.sleep(3)

        current_url = page.url

        # Ürün detay sayfasına yönlendirildiyse
        if '/products/detail/' in current_url:
            result['url'] = current_url
            return _parse_product_page(page, result, qty)

        # Arama sonuçları sayfasındayız - ilk sonuca tıkla
        try:
            # Arama sonuçlarında ürün linki bul
            product_link = page.locator('a[href*="/products/detail/"]').first
            product_link.wait_for(timeout=10000)
            product_link.click()
            time.sleep(3)
            result['url'] = page.url
            return _parse_product_page(page, result, qty)
        except Exception:
            result['status'] = 'NOT_FOUND'
            return result

    except Exception as e:
        result['status'] = f'ERROR: {str(e)[:100]}'
        return result


def _parse_product_page(page, result: dict, qty: int):
    """Ürün detay sayfasından fiyat ve stok bilgilerini parse eder."""
    try:
        # Stok bilgisi - birden fazla format destekle
        try:
            stock_el = page.locator('[data-testid="title-messages"]').first
            stock_text = stock_el.inner_text(timeout=5000)
            # "In-Stock: 5,253,771" formatı
            stock_match = re.search(r'In-Stock:\s*([\d,]+)', stock_text)
            if stock_match:
                result['stock'] = int(stock_match.group(1).replace(',', ''))
            elif 'Available To Order' in stock_text:
                result['stock'] = 'Available To Order'
            elif 'Out of Stock' in stock_text:
                result['stock'] = 0
        except Exception:
            # Fallback: eski yöntem
            try:
                stock_el = page.locator('text=/In-Stock:/').first
                stock_text = stock_el.inner_text(timeout=3000)
                stock_match = re.search(r'In-Stock:\s*([\d,]+)', stock_text)
                if stock_match:
                    result['stock'] = int(stock_match.group(1).replace(',', ''))
            except Exception:
                result['stock'] = None

        # Fiyat: Sadece "Cut Tape (CT) & Digi-Reel®" başlığı altındaki tabloyu kullan
        try:
            pricing_container = page.locator('[data-testid="pricing-table-container"]')
            pricing_container.wait_for(timeout=5000)

            # pricing-group span'larını bul - her biri bir paketleme tipi (Cut Tape, Tape & Reel, Bulk)
            pricing_groups = pricing_container.locator('span[data-testid="pricing-group"]').all()

            ct_table = None  # Cut Tape tablosu
            bulk_table = None  # Bulk tablosu (fallback)

            for group in pricing_groups:
                try:
                    group_text = group.inner_text(timeout=3000)
                    table = group.locator('table').first
                    if 'Cut Tape' in group_text or 'Digi-Reel' in group_text:
                        ct_table = table
                        break
                    elif 'Bulk' in group_text and bulk_table is None:
                        bulk_table = table
                except Exception:
                    continue

            # Cut Tape bulunamazsa Bulk'u kullan
            target_table = ct_table if ct_table else bulk_table

            # Eğer pricing-group yapısı yoksa, container'daki ilk tabloyu al
            if target_table is None:
                all_tables = pricing_container.locator('table').all()
                if all_tables:
                    target_table = all_tables[0]

            best_price = None
            best_qty_break = -1  # En büyük kırılım noktasını takip et
            first_price = None

            if target_table:
                rows = target_table.locator('tbody tr').all()
                for row in rows:
                    cells = row.locator('td').all()
                    if len(cells) >= 2:
                        qty_text = cells[0].inner_text(timeout=2000).strip()
                        price_text = cells[1].inner_text(timeout=2000).strip()

                        qty_val = parse_price_from_text(qty_text)
                        price_val = parse_price_from_text(price_text)

                        if qty_val is not None and price_val is not None:
                            if first_price is None:
                                first_price = price_val
                            # Sipariş adedimize uygun en büyük kırılım noktasının fiyatını al
                            # Örn: adet=4100, kırılımlar 1→$0.08, 10→$0.016
                            # 10 <= 4100 ve 10 > 1 → best = $0.016
                            if qty_val <= qty and qty_val > best_qty_break:
                                best_qty_break = qty_val
                                best_price = price_val

            result['unit_price'] = first_price
            result['qty_price'] = best_price if best_price else first_price

        except Exception:
            pass

        # Description
        try:
            desc_el = page.locator('td:has-text("Description") + td').first
            result['description'] = desc_el.inner_text(timeout=3000).strip()
        except Exception:
            pass

        result['status'] = 'OK' if result['qty_price'] else 'NO_PRICE'
        return result

    except Exception as e:
        result['status'] = f'PARSE_ERROR: {str(e)[:80]}'
        return result


def build_excel_report(results: list, output_dir: str, source_file: str):
    """Sonuçlardan Excel raporu oluşturur."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"digikey_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Digikey Sonuçları"

    # Stiller
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        "Ürün Kodu", "Adet", "Birim Fiyat (USD)", "Adet Fiyatı (USD)",
        "Toplam Maliyet (USD)", "Stok", "Durum", "Açıklama", "URL"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, r in enumerate(results, 2):
        qty = r['qty']
        unit_price = r['qty_price']
        total = unit_price * qty if unit_price else None

        row_data = [
            r['code'], qty, r['unit_price'], unit_price,
            total, r['stock'], r['status'], r['description'], r['url']
        ]

        row_fill = ok_fill if r['status'] == 'OK' else err_fill

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if r['status'] != 'OK':
                cell.fill = row_fill

        # Fiyat formatı
        for col in [3, 4, 5]:
            ws.cell(row=i, column=col).number_format = '#,##0.00000'

    # Sütun genişlikleri
    widths = [25, 10, 18, 18, 20, 15, 15, 40, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Autofilter
    ws.auto_filter.ref = f"A1:I{len(results) + 1}"

    # Freeze
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def load_previous_results(excel_path: str):
    """Önceki rapor Excel'inden sonuçları yükler."""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        results.append({
            'code': str(row[0]).strip(),
            'qty': int(row[1]) if row[1] else 0,
            'unit_price': row[2],
            'qty_price': row[3],
            'stock': row[5],
            'status': str(row[6]).strip() if row[6] else 'ERROR',
            'description': str(row[7]).strip() if row[7] else '',
            'url': str(row[8]).strip() if row[8] else '',
        })
    wb.close()
    return results


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--retry', help='Güncellenecek rapor Excel dosyası')
    parser.add_argument('--retry-codes', help='Tekrar taranacak ürün kodları listesi (dosya yolu, her satırda bir kod)')
    parser.add_argument('--retry-status', default='NO_PRICE', help='Tekrar taranacak durum filtresi (varsayılan: NO_PRICE)')
    args = parser.parse_args()

    if args.retry:
        all_results = load_previous_results(args.retry)

        if args.retry_codes:
            # Dosyadan kod listesi oku
            with open(args.retry_codes) as f:
                retry_set = {line.strip() for line in f if line.strip()}
            retry_indices = [i for i, r in enumerate(all_results) if r['code'] in retry_set]
            print(f"Güncel rapor: {len(all_results)} ürün, {len(retry_indices)} tanesi tekrar taranacak (kod listesinden).")
        else:
            retry_indices = [i for i, r in enumerate(all_results) if r['status'] == args.retry_status]
            print(f"Güncel rapor: {len(all_results)} ürün, {len(retry_indices)} tanesi {args.retry_status} - tekrar taranacak.")

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                viewport={'width': 1280, 'height': 800},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
            )
            page = context.new_page()

            for count, idx in enumerate(retry_indices, 1):
                r = all_results[idx]
                code, qty = r['code'], r['qty']
                print(f"[{count}/{len(retry_indices)}] {code} (adet: {qty})...", end=" ", flush=True)
                new_result = scrape_digikey_product(page, code, qty)
                # Sonucu güncelle
                all_results[idx] = new_result

                price_str = f"${new_result['qty_price']:.5f}" if new_result['qty_price'] else "-"
                stock_val = new_result['stock']
                if isinstance(stock_val, int):
                    stock_str = f"{stock_val:,}"
                elif stock_val:
                    stock_str = str(stock_val)
                else:
                    stock_str = "-"
                print(f"Fiyat: {price_str} | Stok: {stock_str} | {new_result['status']}")

                time.sleep(1.5)

            browser.close()

        results = all_results
    else:
        # Tam tarama
        products = read_products("oguz.xlsx")
        print(f"Toplam {len(products)} ürün okundu.")
        results = []

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                viewport={'width': 1280, 'height': 800},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
            )
            page = context.new_page()

            for i, (code, qty) in enumerate(products):
                print(f"[{i+1}/{len(products)}] {code} (adet: {qty})...", end=" ", flush=True)
                result = scrape_digikey_product(page, code, qty)
                results.append(result)

                price_str = f"${result['qty_price']:.5f}" if result['qty_price'] else "-"
                stock_val = result['stock']
                if isinstance(stock_val, int):
                    stock_str = f"{stock_val:,}"
                elif stock_val:
                    stock_str = str(stock_val)
                else:
                    stock_str = "-"
                print(f"Fiyat: {price_str} | Stok: {stock_str} | {result['status']}")

                time.sleep(1.5)

            browser.close()

    # Rapor oluştur
    output_path = build_excel_report(results, "output", "oguz.xlsx")
    print(f"\nRapor oluşturuldu: {output_path}")
    print(f"Toplam: {len(results)} ürün")
    ok_count = sum(1 for r in results if r['status'] == 'OK')
    print(f"Başarılı: {ok_count}, Başarısız: {len(results) - ok_count}")


if __name__ == "__main__":
    main()
