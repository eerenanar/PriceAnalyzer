"""Excel rapor oluşturucu modülü."""

import logging
import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .scraper import PriceResult

logger = logging.getLogger(__name__)

# Renk sabitleri
COLOR_HEADER_BG = "1F4E79"   # Koyu mavi
COLOR_HEADER_FG = "FFFFFF"   # Beyaz
COLOR_CHEAPEST  = "E2EFDA"   # Açık yeşil
COLOR_ODD_ROW   = "F2F2F2"   # Açık gri
COLOR_ERROR     = "FCE4D6"   # Açık turuncu
COLOR_TOTAL     = "FFF2CC"   # Açık sarı (toplam hücreleri)


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel sheet adı için geçersiz karakterleri temizler."""
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:max_len]


def _apply_header_style(cell, bold: bool = True):
    cell.font = Font(bold=bold, color=COLOR_HEADER_FG)
    cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _find_cheapest(results: List[PriceResult], qty: int = 1) -> Optional[PriceResult]:
    """OK durumundaki sonuçlar arasından verilen adete göre en ucuz olanı döndürür."""
    valid = [r for r in results if r.status == "OK" and r.price is not None]
    if not valid:
        return None
    return min(valid, key=lambda r: r.price_for_qty(qty) or float("inf"))



def build_report(
    all_results: Dict[str, List[PriceResult]],
    output_dir: str,
    offer_quantities: Optional[Dict[str, Tuple[int, int]]] = None,
    input_file: str = "",
) -> str:
    """
    Tüm ürünler için Excel raporu oluşturur.

    Args:
        all_results: {ürün_kodu: [PriceResult, ...]} sözlüğü
        output_dir: Çıktı klasörü
        offer_quantities: {ürün_kodu: (teklif1_adet, teklif2_adet)} sözlüğü
        input_file: Girdi dosyası adı (output dosya adında kullanılır)

    Returns:
        Oluşturulan Excel dosyasının yolu
    """
    if offer_quantities is None:
        offer_quantities = {}

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(input_file))[0] if input_file else "output"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"{base_name}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Teklif-1 En Ucuz ─────────────────────────────────────────────
    ws_t1 = wb.active
    ws_t1.title = "Teklif-1 En Ucuz"

    t1_headers = [
        "Ürün Kodu", "Teklif-1 Adet",
        "En Ucuz Birim Fiyat", "Para Birimi", "Teklif-1 Toplam",
        "Stok Adedi", "Kaynak Site", "URL", "Tarama Tarihi",
    ]
    ws_t1.append(t1_headers)
    for col_idx, _ in enumerate(t1_headers, 1):
        _apply_header_style(ws_t1.cell(row=1, column=col_idx))
    ws_t1.row_dimensions[1].height = 25

    for row_idx, (product_code, results) in enumerate(all_results.items(), start=2):
        qty1, _ = offer_quantities.get(product_code, (0, 0))
        cheapest1 = _find_cheapest(results, qty=qty1 if qty1 > 0 else 1)

        if qty1 == 0:
            row_data = [product_code, "-", "-", "", "", "", "", "", ""]
            fill_color = None
        elif cheapest1:
            unit_price = cheapest1.price_for_qty(qty1)
            total = round(unit_price * qty1, 2) if unit_price else None
            stock_display = (
                "Stok Yok" if cheapest1.stock == 0
                else ("Var" if cheapest1.stock == -1 else str(cheapest1.stock))
            )
            row_data = [
                product_code, qty1,
                unit_price, cheapest1.currency, total,
                stock_display, cheapest1.site_name, cheapest1.url, cheapest1.scraped_at,
            ]
            fill_color = COLOR_CHEAPEST
        else:
            row_data = [product_code, qty1, "Bulunamadı", "", "", "", "", "", ""]
            fill_color = COLOR_ERROR

        ws_t1.append(row_data)
        for col_idx in range(1, len(t1_headers) + 1):
            cell = ws_t1.cell(row=row_idx, column=col_idx)
            if col_idx == 5 and isinstance(cell.value, (int, float)):
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)
            elif fill_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for price_col in (3, 5):
        for row in ws_t1.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00000'

    _auto_fit_columns(ws_t1)
    ws_t1.freeze_panes = "A2"

    # ── Sheet 2: Teklif-2 En Ucuz ─────────────────────────────────────────────
    ws_t2 = wb.create_sheet(title="Teklif-2 En Ucuz")

    t2_headers = [
        "Ürün Kodu", "Teklif-2 Adet",
        "En Ucuz Birim Fiyat", "Para Birimi", "Teklif-2 Toplam",
        "Stok Adedi", "Kaynak Site", "URL", "Tarama Tarihi",
    ]
    ws_t2.append(t2_headers)
    for col_idx, _ in enumerate(t2_headers, 1):
        _apply_header_style(ws_t2.cell(row=1, column=col_idx))
    ws_t2.row_dimensions[1].height = 25

    for row_idx, (product_code, results) in enumerate(all_results.items(), start=2):
        _, qty2 = offer_quantities.get(product_code, (0, 0))
        cheapest2 = _find_cheapest(results, qty=qty2 if qty2 > 0 else 1)

        if qty2 == 0:
            row_data = [product_code, "-", "-", "", "", "", "", "", ""]
            fill_color = None
        elif cheapest2:
            unit_price = cheapest2.price_for_qty(qty2)
            total = round(unit_price * qty2, 2) if unit_price else None
            stock_display = (
                "Stok Yok" if cheapest2.stock == 0
                else ("Var" if cheapest2.stock == -1 else str(cheapest2.stock))
            )
            row_data = [
                product_code, qty2,
                unit_price, cheapest2.currency, total,
                stock_display, cheapest2.site_name, cheapest2.url, cheapest2.scraped_at,
            ]
            fill_color = COLOR_CHEAPEST
        else:
            row_data = [product_code, qty2, "Bulunamadı", "", "", "", "", "", ""]
            fill_color = COLOR_ERROR

        ws_t2.append(row_data)
        for col_idx in range(1, len(t2_headers) + 1):
            cell = ws_t2.cell(row=row_idx, column=col_idx)
            if col_idx == 5 and isinstance(cell.value, (int, float)):
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)
            elif fill_color:
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for price_col in (3, 5):
        for row in ws_t2.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00000'

    _auto_fit_columns(ws_t2)
    ws_t2.freeze_panes = "A2"

    # ── Sheet 3: Genel Özet (Teklif-1 + Teklif-2 birleşik) ───────────────────
    ws_ozet = wb.create_sheet(title="Genel Özet")

    ozet_headers = [
        "Ürün Kodu",
        "Teklif-1 Adet", "T1 Birim Fiyat", "T1 Para Birimi", "T1 Toplam", "T1 Kaynak", "T1 URL",
        "Teklif-2 Adet", "T2 Birim Fiyat", "T2 Para Birimi", "T2 Toplam", "T2 Kaynak", "T2 URL",
    ]
    ws_ozet.append(ozet_headers)
    for col_idx, _ in enumerate(ozet_headers, 1):
        _apply_header_style(ws_ozet.cell(row=1, column=col_idx))
    ws_ozet.row_dimensions[1].height = 25

    for row_idx, (product_code, results) in enumerate(all_results.items(), start=2):
        qty1, qty2 = offer_quantities.get(product_code, (0, 0))
        cheapest1 = _find_cheapest(results, qty=qty1 if qty1 > 0 else 1)
        cheapest2 = _find_cheapest(results, qty=qty2 if qty2 > 0 else 1)

        # Teklif-1 sütunları
        if qty1 == 0:
            t1_cols = ["-", "-", "", "", "", ""]
            t1_url = ""
            t1_fill = None
        elif cheapest1:
            up1 = cheapest1.price_for_qty(qty1)
            t1_url = cheapest1.url
            t1_cols = [qty1, up1, cheapest1.currency, round(up1 * qty1, 2) if up1 else None, cheapest1.site_name, t1_url]
            t1_fill = COLOR_CHEAPEST
        else:
            t1_cols = [qty1, "Bulunamadı", "", "", "", ""]
            t1_url = ""
            t1_fill = COLOR_ERROR

        # Teklif-2 sütunları
        if qty2 == 0:
            t2_cols = ["-", "-", "", "", "", ""]
            t2_url = ""
            t2_fill = None
        elif cheapest2:
            up2 = cheapest2.price_for_qty(qty2)
            t2_url = cheapest2.url
            t2_cols = [qty2, up2, cheapest2.currency, round(up2 * qty2, 2) if up2 else None, cheapest2.site_name, t2_url]
            t2_fill = COLOR_CHEAPEST
        else:
            t2_cols = [qty2, "Bulunamadı", "", "", "", ""]
            t2_url = ""
            t2_fill = COLOR_ERROR

        ws_ozet.append([product_code] + t1_cols + t2_cols)

        for col_idx in range(1, len(ozet_headers) + 1):
            cell = ws_ozet.cell(row=row_idx, column=col_idx)
            # T1 toplam = col 5, T2 toplam = col 11
            if col_idx in (5, 11) and isinstance(cell.value, (int, float)):
                cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)
            elif col_idx in (2, 3, 4, 5, 6, 7) and t1_fill:
                cell.fill = PatternFill(fill_type="solid", fgColor=t1_fill)
            elif col_idx in (8, 9, 10, 11, 12, 13) and t2_fill:
                cell.fill = PatternFill(fill_type="solid", fgColor=t2_fill)
            # URL hücrelerine hyperlink ekle
            if col_idx == 7 and t1_url:
                cell.hyperlink = t1_url
                cell.font = Font(color="0563C1", underline="single")
            elif col_idx == 13 and t2_url:
                cell.hyperlink = t2_url
                cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for price_col in (3, 5, 9, 11):
        for row in ws_ozet.iter_rows(min_row=2, min_col=price_col, max_col=price_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00000'

    _auto_fit_columns(ws_ozet)
    ws_ozet.freeze_panes = "A2"

    # ── Per-Ürün Sheetleri ─────────────────────────────────────────────────────
    detail_headers = [
        "Site Adı", "Birim Fiyat", "Para Birimi",
        "Stok Adedi", "URL", "Durum", "Tarama Tarihi",
    ]

    for product_code, results in all_results.items():
        sheet_name = _safe_sheet_name(product_code)
        ws = wb.create_sheet(title=sheet_name)

        qty1, qty2 = offer_quantities.get(product_code, (0, 0))

        # Teklif bilgisini sheet başına yaz
        ws.append([f"Ürün: {product_code}",
                   f"Teklif-1 Adet: {qty1}",
                   f"Teklif-2 Adet: {qty2}"])
        ws.row_dimensions[1].height = 20

        # Teklif bazlı ek sütunları belirle
        extra_headers = []
        if qty1 > 0:
            extra_headers += [f"Birim Fiyat (T1:{qty1})", f"Teklif-1 Toplam"]
        if qty2 > 0:
            extra_headers += [f"Birim Fiyat (T2:{qty2})", f"Teklif-2 Toplam"]

        ws.append(detail_headers + extra_headers)
        header_row_idx = 2
        for col_idx, _ in enumerate(detail_headers + extra_headers, 1):
            _apply_header_style(ws.cell(row=header_row_idx, column=col_idx))
        ws.row_dimensions[header_row_idx].height = 25

        cheapest1 = _find_cheapest(results, qty=qty1 if qty1 > 0 else 1)
        cheapest2 = _find_cheapest(results, qty=qty2 if qty2 > 0 else 1)

        for row_idx, result in enumerate(results, start=3):
            stock_display = (
                "Stok Yok" if result.stock == 0
                else ("Var" if result.stock == -1 else (str(result.stock) if result.stock is not None else ""))
            )
            # Birim fiyat: 1+ (liste başı) fiyatı
            unit_price = result.price_for_qty(1) if result.price is not None else None
            row_data = [
                result.site_name,
                unit_price,
                result.currency,
                stock_display,
                result.url,
                result.status,
                result.scraped_at,
            ]
            if qty1 > 0:
                p1 = result.price_for_qty(qty1) if result.price is not None else None
                total1 = round(p1 * qty1, 2) if p1 else None
                row_data += [p1, total1]
            if qty2 > 0:
                p2 = result.price_for_qty(qty2) if result.price is not None else None
                total2 = round(p2 * qty2, 2) if p2 else None
                row_data += [p2, total2]

            ws.append(row_data)

            # Satır rengi: her iki teklif için de en ucuz vurgulansın
            is_cheapest = (
                (cheapest1 and result is cheapest1) or
                (cheapest2 and result is cheapest2)
            )
            if is_cheapest:
                bg = COLOR_CHEAPEST
            elif result.status != "OK":
                bg = COLOR_ERROR
            elif row_idx % 2 == 0:
                bg = COLOR_ODD_ROW
            else:
                bg = None

            total_cols = len(detail_headers) + len(extra_headers)
            extra_start = len(detail_headers) + 1
            for col_idx in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Toplam sütunları (çift indexli extra kolonlar) sarı
                extra_offset = col_idx - extra_start
                if col_idx >= extra_start and extra_offset % 2 == 1 and cell.value is not None:
                    cell.fill = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL)
                elif bg:
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Fiyat formatı: birim fiyat (col 2) + tüm extra sayısal sütunlar
        price_cols = [2] + list(range(len(detail_headers) + 1, len(detail_headers) + len(extra_headers) + 1))
        for price_col in price_cols:
            for row in ws.iter_rows(min_row=3, min_col=price_col, max_col=price_col):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00000'

        _auto_fit_columns(ws)
        ws.freeze_panes = "A3"

    wb.save(output_path)
    logger.info(f"Rapor oluşturuldu: {output_path}")
    return output_path
